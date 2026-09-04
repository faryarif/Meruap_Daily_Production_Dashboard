"""Monthly production and lifting workbook ETL, persistence, and charts."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook
from plotly.subplots import make_subplots
from supabase import create_client


MONTHLY_COLUMNS = [
    "report_month",
    "reporting_status",
    "source_filename",
    "rc_bopd",
    "rc_volume_bbl",
    "rc_pumping_net_bbl",
    "rc_pumping_gross_bbl",
    "rc_received_bbl",
    "field_production_bbl",
    "bsa_production_bbl",
    "bsa_transfer_bbl",
    "bsa_storage_loss_gain_bbl",
    "bsa_stock_movement_bbl",
    "bsb_production_bbl",
    "bsb_transfer_bbl",
    "bsb_storage_loss_gain_bbl",
    "bsb_stock_movement_bbl",
    "sta_received_bbl",
    "sta_transfer_bbl",
    "sta_storage_loss_gain_bbl",
    "sta_stock_movement_bbl",
    "bajubang_received_bbl",
    "bajubang_pumped_bbl",
    "bajubang_storage_loss_gain_bbl",
    "bajubang_stock_movement_bbl",
    "shipping_received_bbl",
    "tempino_opening_stock_bbl",
    "tempino_closing_stock_bbl",
    "tempino_meter_gross_bbl",
    "tempino_storage_loss_gain_bbl",
    "tempino_pumping_net_bbl",
    "s_gerong_pumped_bbl",
    "s_gerong_received_bbl",
]

NUMERIC_MONTHLY_COLUMNS = [c for c in MONTHLY_COLUMNS if c.endswith("_bbl") or c == "rc_bopd"]

ROW_MAP = {
    "field_production_bbl": 6,
    "bsa_production_bbl": 10,
    "bsa_transfer_bbl": 11,
    "bsa_storage_loss_gain_bbl": 13,
    "bsa_stock_movement_bbl": 14,
    "bsb_production_bbl": 19,
    "bsb_transfer_bbl": 20,
    "bsb_storage_loss_gain_bbl": 23,
    "bsb_stock_movement_bbl": 24,
    "sta_received_bbl": 29,
    "sta_transfer_bbl": 30,
    "sta_storage_loss_gain_bbl": 33,
    "sta_stock_movement_bbl": 35,
    "bajubang_received_bbl": 40,
    "bajubang_pumped_bbl": 41,
    "bajubang_storage_loss_gain_bbl": 43,
    "bajubang_stock_movement_bbl": 44,
    "shipping_received_bbl": 46,
    "tempino_opening_stock_bbl": 51,
    "tempino_closing_stock_bbl": 54,
    "tempino_meter_gross_bbl": 53,
    "tempino_storage_loss_gain_bbl": 56,
    "tempino_pumping_net_bbl": 58,
    "s_gerong_pumped_bbl": 142,
    "s_gerong_received_bbl": 143,
}


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, str):
        value = value.replace(",", "").strip()
    result = pd.to_numeric(value, errors="coerce")
    return None if pd.isna(result) else float(result)


class _WorkbookReader:
    sheet_names: list[str]

    def value(self, sheet_name: str, row: int, column: int) -> Any:
        raise NotImplementedError

    def max_column(self, sheet_name: str) -> int:
        raise NotImplementedError

    def close(self) -> None:
        raise NotImplementedError


class _XlsxReader(_WorkbookReader):
    def __init__(self, raw_bytes: bytes):
        self.book = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        self.sheet_names = list(self.book.sheetnames)

    def value(self, sheet_name: str, row: int, column: int) -> Any:
        return self.book[sheet_name].cell(row=row, column=column).value

    def max_column(self, sheet_name: str) -> int:
        return self.book[sheet_name].max_column

    def close(self) -> None:
        self.book.close()


class _XlsReader(_WorkbookReader):
    def __init__(self, raw_bytes: bytes):
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("Reading .xls files requires xlrd 2.0.2.") from exc
        self.book = xlrd.open_workbook(file_contents=raw_bytes, on_demand=True)
        self.sheet_names = list(self.book.sheet_names())

    def value(self, sheet_name: str, row: int, column: int) -> Any:
        sheet = self.book.sheet_by_name(sheet_name)
        if row > sheet.nrows or column > sheet.ncols:
            return None
        return sheet.cell_value(row - 1, column - 1)

    def max_column(self, sheet_name: str) -> int:
        return self.book.sheet_by_name(sheet_name).ncols

    def close(self) -> None:
        self.book.release_resources()


def _open_reader(raw_bytes: bytes, filename: str) -> _WorkbookReader:
    suffix = str(filename).lower().rsplit(".", 1)[-1]
    if suffix == "xls":
        return _XlsReader(raw_bytes)
    if suffix == "xlsx":
        return _XlsxReader(raw_bytes)
    raise ValueError("Only .xls and .xlsx workbooks are supported.")


def _read_rc_rows(reader: _WorkbookReader, year: int) -> dict[int, dict[str, float | None]]:
    sheet_name = f"RC {year}"
    if sheet_name not in reader.sheet_names:
        return {}
    rows: dict[int, dict[str, float | None]] = {}
    for row in range(3, 15):
        month_number = row - 2
        rows[month_number] = {
            "rc_bopd": _number(reader.value(sheet_name, row, 3)),
            "rc_volume_bbl": _number(reader.value(sheet_name, row, 4)),
            "rc_pumping_net_bbl": _number(reader.value(sheet_name, row, 5)),
            "rc_pumping_gross_bbl": _number(reader.value(sheet_name, row, 6)),
            "rc_received_bbl": _number(reader.value(sheet_name, row, 7)),
        }
    return rows


def parse_lifting_workbook(uploaded_file) -> tuple[pd.DataFrame, list[str]]:
    """Extract every recognized yearly sheet and its matching RC sheet."""
    raw_bytes = uploaded_file.getvalue()
    filename = getattr(uploaded_file, "name", "monthly_lifting.xls")
    reader = _open_reader(raw_bytes, filename)
    warnings: list[str] = []
    records: list[dict[str, Any]] = []
    try:
        years = sorted(
            int(name)
            for name in reader.sheet_names
            if re.fullmatch(r"20\d{2}", str(name)) and f"RC {name}" in reader.sheet_names
        )
        if not years:
            raise ValueError("No yearly sheet with a matching 'RC YYYY' sheet was found.")
        latest_year = max(years)
        for year in years:
            sheet_name = str(year)
            rc_rows = _read_rc_rows(reader, year)
            month_columns: dict[int, int] = {}
            for column in range(4, reader.max_column(sheet_name) + 1):
                month_value = _number(reader.value(sheet_name, 5, column))
                if month_value is not None and month_value.is_integer() and 1 <= int(month_value) <= 12:
                    month_columns[int(month_value)] = column

            for month_number, column in sorted(month_columns.items()):
                record: dict[str, Any] = {
                    "report_month": date(year, month_number, 1).isoformat(),
                    "source_filename": filename,
                }
                record.update(rc_rows.get(month_number, {}))
                for target, row in ROW_MAP.items():
                    record[target] = _number(reader.value(sheet_name, row, column))

                rc_received = record.get("rc_received_bbl") or 0.0
                sheet_received = record.get("s_gerong_received_bbl") or 0.0
                if rc_received > 0:
                    record["s_gerong_received_bbl"] = rc_received
                if (record.get("rc_pumping_net_bbl") or 0.0) > 0:
                    record["s_gerong_pumped_bbl"] = record["rc_pumping_net_bbl"]

                has_actual = (record.get("field_production_bbl") or 0.0) > 0 and (
                    record.get("s_gerong_received_bbl") or 0.0
                ) > 0
                has_reference = (record.get("rc_volume_bbl") or 0.0) > 0
                if has_actual:
                    record["reporting_status"] = "Actual"
                elif year == latest_year and has_reference:
                    record["reporting_status"] = "Planned"
                else:
                    record["reporting_status"] = "Missing"

                if rc_received > 0 and sheet_received > 0 and abs(rc_received - sheet_received) > 0.1:
                    warnings.append(
                        f"{year}-{month_number:02d}: RC received and the yearly-sheet received value differ by "
                        f"{abs(rc_received - sheet_received):,.1f} bbl. RC received is used as official lifting."
                    )
                records.append(record)
    finally:
        reader.close()

    result = pd.DataFrame(records).reindex(columns=MONTHLY_COLUMNS)
    return normalize_monthly_data(result), warnings


def normalize_monthly_data(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=MONTHLY_COLUMNS)
    normalized = df.copy()
    for column in MONTHLY_COLUMNS:
        if column not in normalized.columns:
            normalized[column] = None
    normalized["report_month"] = pd.to_datetime(normalized["report_month"], errors="coerce")
    for column in NUMERIC_MONTHLY_COLUMNS:
        normalized[column] = pd.to_numeric(normalized[column], errors="coerce")
    normalized["reporting_status"] = normalized["reporting_status"].fillna("Missing")
    return normalized[MONTHLY_COLUMNS].sort_values("report_month").reset_index(drop=True)


@st.cache_data(ttl=300, show_spinner=False)
def read_monthly_reconciliation(start_month: str | None = None, end_month: str | None = None) -> pd.DataFrame:
    params = {}
    if start_month:
        params["p_start_month"] = start_month
    if end_month:
        params["p_end_month"] = end_month
    response = _admin_client().rpc("dashboard_monthly_reconciliation", params).execute()
    return normalize_monthly_data(pd.DataFrame(response.data or []))


def _admin_client():
    """Server-side client; the service key must only live in Streamlit secrets."""
    key = st.secrets["supabase"]["service_role_key"]
    return create_client(st.secrets["supabase"]["url"].rstrip("/"), key)


def upload_monthly_reconciliation(df: pd.DataFrame) -> int:
    normalized = normalize_monthly_data(df)
    if normalized.empty:
        return 0
    records = []
    for raw in normalized.to_dict(orient="records"):
        record: dict[str, Any] = {}
        for column in MONTHLY_COLUMNS:
            value = raw.get(column)
            if column == "report_month":
                record[column] = pd.Timestamp(value).date().isoformat()
            elif pd.isna(value):
                record[column] = None
            elif column in NUMERIC_MONTHLY_COLUMNS:
                record[column] = float(value)
            else:
                record[column] = str(value)
        records.append(record)

    client = _admin_client()
    for start in range(0, len(records), 500):
        response = client.table("monthly_lifting_reconciliation").upsert(
            records[start : start + 500], on_conflict="report_month"
        ).execute()
        if response.data is None:
            raise RuntimeError("Supabase returned no data while saving monthly reconciliation.")
    read_monthly_reconciliation.clear()
    return len(records)


@dataclass(frozen=True)
class LossSegment:
    segment: str
    sent_bbl: float
    received_bbl: float

    @property
    def loss_bbl(self) -> float:
        return self.sent_bbl - self.received_bbl

    @property
    def loss_pct(self) -> float:
        return self.loss_bbl / self.sent_bbl * 100.0 if self.sent_bbl else np.nan


def _value(row: pd.Series | dict[str, Any], key: str) -> float:
    value = row.get(key, 0.0)
    return 0.0 if pd.isna(value) else float(value)


def loss_segments(row: pd.Series | dict[str, Any]) -> list[LossSegment]:
    block_sent = _value(row, "bsa_transfer_bbl") + _value(row, "bsb_transfer_bbl")
    return [
        LossSegment("Block Stations → STA", block_sent, _value(row, "sta_received_bbl")),
        LossSegment("STA → Bajubang", _value(row, "sta_transfer_bbl"), _value(row, "bajubang_received_bbl")),
        LossSegment(
            "Bajubang → Shipping Tank", _value(row, "bajubang_pumped_bbl"), _value(row, "shipping_received_bbl")
        ),
        LossSegment(
            "Shipping Tank → Tempino Meter",
            _value(row, "tempino_meter_gross_bbl"),
            _value(row, "tempino_pumping_net_bbl"),
        ),
        LossSegment(
            "Tempino → S. Gerong", _value(row, "s_gerong_pumped_bbl"), _value(row, "s_gerong_received_bbl")
        ),
    ]


def heatmap_loss_segments(row: pd.Series | dict[str, Any]) -> list[LossSegment]:
    """Return transfer losses in physical field-to-lifting order."""
    block_production = _value(row, "bsa_production_bbl") + _value(row, "bsb_production_bbl")
    block_transfer = _value(row, "bsa_transfer_bbl") + _value(row, "bsb_transfer_bbl")
    return [
        LossSegment("Production at Field → Block Station A+B", _value(row, "field_production_bbl"), block_production),
        LossSegment("Block Station A+B → Staging Area", block_transfer, _value(row, "sta_received_bbl")),
        LossSegment("Staging Area → SPU Bajubang", _value(row, "sta_transfer_bbl"), _value(row, "bajubang_received_bbl")),
        LossSegment("SPU Bajubang → PPP Tempino", _value(row, "bajubang_pumped_bbl"), _value(row, "tempino_meter_gross_bbl")),
        LossSegment("PPP Tempino → S. Gerong", _value(row, "tempino_pumping_net_bbl"), _value(row, "s_gerong_received_bbl")),
    ]


def loss_segment_frame(row: pd.Series | dict[str, Any]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Segment": item.segment,
                "Sent (bbl)": item.sent_bbl,
                "Received (bbl)": item.received_bbl,
                "Magnitude (bbl)": abs(item.loss_bbl),
                "Magnitude (%)": abs(item.loss_pct),
                "Signed Difference (bbl)": item.loss_bbl,
                "Result": "Loss" if item.loss_bbl > 0 else "Gain" if item.loss_bbl < 0 else "Balanced",
            }
            for item in loss_segments(row)
        ]
    )


def reconciliation_components(row: pd.Series | dict[str, Any]) -> list[tuple[str, float]]:
    components = [
        ("Block station stock movement", _value(row, "bsa_stock_movement_bbl") + _value(row, "bsb_stock_movement_bbl")),
        (
            "Block station storage loss/gain",
            -(_value(row, "bsa_storage_loss_gain_bbl") + _value(row, "bsb_storage_loss_gain_bbl")),
        ),
        (
            "Block Stations → STA",
            _value(row, "bsa_transfer_bbl") + _value(row, "bsb_transfer_bbl") - _value(row, "sta_received_bbl"),
        ),
        ("STA stock movement", _value(row, "sta_stock_movement_bbl")),
        ("STA storage loss/gain", -_value(row, "sta_storage_loss_gain_bbl")),
        ("STA → Bajubang", _value(row, "sta_transfer_bbl") - _value(row, "bajubang_received_bbl")),
        ("Bajubang stock movement", _value(row, "bajubang_stock_movement_bbl")),
        ("Bajubang storage loss/gain", -_value(row, "bajubang_storage_loss_gain_bbl")),
        (
            "Bajubang → Shipping Tank",
            _value(row, "bajubang_pumped_bbl") - _value(row, "shipping_received_bbl"),
        ),
        (
            "Tempino stock movement",
            _value(row, "tempino_closing_stock_bbl") - _value(row, "tempino_opening_stock_bbl"),
        ),
        ("Tempino storage loss/gain", -_value(row, "tempino_storage_loss_gain_bbl")),
        (
            "Shipping Tank → Tempino Meter",
            _value(row, "tempino_meter_gross_bbl") - _value(row, "tempino_pumping_net_bbl"),
        ),
        (
            "Tempino → S. Gerong",
            _value(row, "s_gerong_pumped_bbl") - _value(row, "s_gerong_received_bbl"),
        ),
    ]
    gap = _value(row, "field_production_bbl") - _value(row, "s_gerong_received_bbl")
    residual = gap - sum(value for _, value in components)
    components.append(("Unexplained balance", residual))
    return components


def make_monthly_trend_figure(df: pd.DataFrame) -> go.Figure:
    actual = normalize_monthly_data(df)
    actual = actual[actual["reporting_status"].eq("Actual")]
    fig = go.Figure()
    series = [
        ("Field Production", "field_production_bbl", "#22c55e"),
        ("RC Actual Reference", "rc_volume_bbl", "#f59e0b"),
        ("Pumping Net", "rc_pumping_net_bbl", "#38bdf8"),
        ("Received S. Gerong", "s_gerong_received_bbl", "#a855f7"),
    ]
    for label, column, color in series:
        fig.add_trace(
            go.Scatter(
                x=actual["report_month"],
                y=actual[column],
                name=label,
                mode="lines",
                line={"color": color, "width": 2.5},
                hovertemplate=f"{label}<br>%{{x|%b %Y}}<br>%{{y:,.1f}} bbl<extra></extra>",
            )
        )
    return _chart_layout(fig, "Monthly Production and Lifting", "bbl/month")


def make_rc_chart(df: pd.DataFrame) -> go.Figure:
    actual = normalize_monthly_data(df)
    actual = actual[actual["reporting_status"].eq("Actual")]
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Bar(
            x=actual["report_month"],
            y=actual["rc_volume_bbl"],
            name="RC Actual Volume",
            marker_color="#f59e0b",
            hovertemplate="%{x|%b %Y}<br>%{y:,.1f} bbl<extra></extra>",
        ),
        secondary_y=False,
    )
    fig.add_trace(
        go.Scatter(
            x=actual["report_month"],
            y=actual["rc_bopd"],
            name="RC Actual BOPD",
            mode="lines",
            line={"color": "#fef08a", "width": 2.5},
            hovertemplate="%{x|%b %Y}<br>%{y:,.1f} BOPD<extra></extra>",
        ),
        secondary_y=True,
    )
    _chart_layout(fig, "RC Actual Reference", "bbl/month")
    fig.update_yaxes(title_text="BOPD", secondary_y=True, gridcolor="rgba(148,163,184,0.08)")
    return fig


def make_waterfall_figure(row: pd.Series | dict[str, Any]) -> go.Figure:
    components = reconciliation_components(row)
    labels = ["Field Production"] + [name for name, _ in components] + ["Received S. Gerong"]
    values = [_value(row, "field_production_bbl")] + [-value for _, value in components] + [0]
    measures = ["absolute"] + ["relative"] * len(components) + ["total"]
    fig = go.Figure(
        go.Waterfall(
            x=labels,
            y=values,
            measure=measures,
            connector={"line": {"color": "#64748b"}},
            increasing={"marker": {"color": "#22c55e"}},
            decreasing={"marker": {"color": "#ef4444"}},
            totals={"marker": {"color": "#a855f7"}},
            text=[f"{_value(row, 'field_production_bbl'):,.1f}"]
            + [f"{value:,.1f}" for _, value in components]
            + [f"{_value(row, 's_gerong_received_bbl'):,.1f}"],
            textposition="outside",
            hovertemplate="%{x}<br>%{y:,.1f} bbl<extra></extra>",
        )
    )
    return _chart_layout(fig, "Field-to-Lifting Reconciliation", "bbl")


def make_sankey_figure(row: pd.Series | dict[str, Any]) -> go.Figure:
    core = [
        ("Field Production", _value(row, "field_production_bbl")),
        ("Block Station Transfer", _value(row, "bsa_transfer_bbl") + _value(row, "bsb_transfer_bbl")),
        ("STA Received", _value(row, "sta_received_bbl")),
        ("STA Transfer", _value(row, "sta_transfer_bbl")),
        ("Bajubang Received", _value(row, "bajubang_received_bbl")),
        ("Bajubang Pumped", _value(row, "bajubang_pumped_bbl")),
        ("Shipping Tank Received", _value(row, "shipping_received_bbl")),
        ("Tempino Meter Gross", _value(row, "tempino_meter_gross_bbl")),
        ("Pumping Net", _value(row, "tempino_pumping_net_bbl")),
        ("Received S. Gerong", _value(row, "s_gerong_received_bbl")),
    ]
    transition_labels = [
        "Block station stock/storage",
        "Block Stations → STA loss/gain",
        "STA stock/storage",
        "STA → Bajubang loss/gain",
        "Bajubang stock/storage",
        "Bajubang → Shipping loss/gain",
        "Tempino stock/storage",
        "Meter difference",
        "Tempino → S. Gerong loss/gain",
    ]
    labels = [f"{name}<br>{np.ceil(value):,.0f} bbl" for name, value in core]
    colors = ["#22c55e"] + ["#38bdf8"] * 8 + ["#a855f7"]
    source: list[int] = []
    target: list[int] = []
    values: list[float] = []
    link_colors: list[str] = []
    for index, adjustment_name in enumerate(transition_labels):
        start_value = max(core[index][1], 0.0)
        end_value = max(core[index + 1][1], 0.0)
        main_value = min(start_value, end_value)
        if main_value > 0:
            source.append(index)
            target.append(index + 1)
            values.append(float(np.ceil(main_value)))
            link_colors.append("rgba(56,189,248,0.38)")
        difference = start_value - end_value
        if abs(difference) >= 0.05:
            adjustment_index = len(labels)
            if difference > 0:
                labels.append(f"{adjustment_name}<br>Loss {np.ceil(difference):,.0f} bbl")
                colors.append("#ef4444")
                source.append(index)
                target.append(adjustment_index)
                values.append(float(np.ceil(difference)))
                link_colors.append("rgba(239,68,68,0.42)")
            else:
                labels.append(f"{adjustment_name}<br>Gain/release {np.ceil(-difference):,.0f} bbl")
                colors.append("#22c55e")
                source.append(adjustment_index)
                target.append(index + 1)
                values.append(float(np.ceil(-difference)))
                link_colors.append("rgba(34,197,94,0.42)")
    fig = go.Figure(
        go.Sankey(
            arrangement="snap",
            valueformat=",.0f",
            valuesuffix=" bbl",
            node={"label": labels, "color": colors, "pad": 15, "thickness": 16, "line": {"width": 0}},
            link={"source": source, "target": target, "value": values, "color": link_colors},
        )
    )
    fig = _chart_layout(fig, "Production and Lifting Flow", "")
    fig.update_layout(height=650)
    return fig


def make_loss_heatmap(df: pd.DataFrame) -> go.Figure:
    actual = normalize_monthly_data(df)
    actual = actual[actual["reporting_status"].eq("Actual")]
    rows = []
    segment_order: list[str] = []
    for _, record in actual.iterrows():
        items = heatmap_loss_segments(record)
        if not segment_order:
            segment_order = [item.segment for item in items]
        for item in items:
            rows.append({"Month": record["report_month"], "Segment": item.segment, "Loss %": item.loss_pct})
    source = pd.DataFrame(rows)
    if source.empty:
        return _chart_layout(go.Figure(), "Transfer Loss Heatmap", "")
    pivot = source.pivot(index="Segment", columns="Month", values="Loss %").reindex(segment_order)
    text = np.where(pd.isna(pivot.values), "", np.vectorize(lambda x: f"{x:.2f}%")(pivot.fillna(0).values))
    fig = go.Figure(
        go.Heatmap(
            z=pivot.values,
            x=[pd.Timestamp(value).strftime("%b %Y") for value in pivot.columns],
            y=pivot.index,
            colorscale=[[0, "#16a34a"], [0.5, "#facc15"], [1, "#dc2626"]],
            zmid=0,
            text=text,
            texttemplate="%{text}",
            hovertemplate="%{y}<br>%{x}<br>%{z:.2f}%<extra></extra>",
            colorbar={"title": "Loss %"},
        )
    )
    return _chart_layout(fig, "Transfer Loss Heatmap", "")


def _chart_layout(fig: go.Figure, title: str, y_title: str) -> go.Figure:
    fig.update_layout(
        title={"text": title, "font": {"size": 18}},
        template="plotly_dark",
        paper_bgcolor="#141d2e",
        plot_bgcolor="#141d2e",
        font={"color": "#e2e8f0"},
        margin={"l": 30, "r": 25, "t": 65, "b": 35},
        hovermode="x unified",
        legend={"orientation": "h", "y": 1.08, "x": 0},
        height=470,
    )
    fig.update_xaxes(gridcolor="rgba(148,163,184,0.10)")
    fig.update_yaxes(title_text=y_title, gridcolor="rgba(148,163,184,0.10)")
    return fig

